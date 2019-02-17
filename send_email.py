import smtplib 
import openpyxl as xl
import sys, time, calendar, datetime
from getpass import getpass
from termcolor import colored
from datetime import datetime
from datetime import timedelta
import socks

#opening excel file containing names and email of recipients
wb=xl.load_workbook('book.xlsx')
sheet=wb.get_sheet_by_name('sheet 1')
candidates={}
#retrieveing names and emails in dictionary candidates
for r in range(2,sheet.get_highest_row()+1):
	email=sheet.cell(row=r, column=2).value
	name=sheet.cell(row=r, column=3).value
	name=name.title()
	candidates[name]=email
socks.setdefaultproxy(socks.SOCKS5, '192.162.0.103', 3128)
socks.wrapmodule(smtplib)
smtpObj=smtplib.SMTP('smtp-mail.outlook.com')
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('example@outlook.com','password')

#body="Subject: Interview time for Assistant Coordinator. \n\nDear %s,\nYour interview for Assistant Coordinator has been scheduled at %s on %s. Please, reach LT complex 5 minutes before your scheduled time.\nBe on time. Good Luck.\n\nISSACC Coordinators\nAshwani Sharma (+91-9457178883)\nAnushka Awasthi" % (name,tm,dt)
#print ('Sending email to %s...' % name)
#sendmailstatus=smtpObj.sendmail('from','to',body)

t=datetime(2018, 1, 20, 16, 00)
endtime=datetime(2018,1,20,18, 00)
sent={}
notsent={}
for name, email in candidates.items():
	dt=t.strftime('%d %B %Y')
	tm=t.strftime('%I:%M %p')
	str(t.date())
	body="Subject: Interview time. \n\nDear %s,\nYour interview has been scheduled at %s on %s. Please, reach 5 minutes before your scheduled time.\nBe on time. Good Luck.\n\n" % (name,tm,dt)
	print ('Sending email to %s...' % name)
	sendmailstatus=smtpObj.sendmail('from','to',body)
	if sendmailstatus == {}:
		print ("Mail sent to %s." % name)
		sent[name]=[t,dt,tm]
		t += timedelta(minutes=5)
	else:
		notsent[name]=email
		print ("Mail not sent")
	if t==endtime:
		t=datetime(2018, 1, 21, 14, 00)
wb=xl.Workbook()
sheet=wb.active
sheet.title='sent time'
sheet['A1']='Name'
sheet['A2']='Date'
sheet['A3']='Time'
sheet['A4']='email'
x=int(2)
for name, t in sent.items():
	sheet[x+'1']=name
	sheet[x2]=t
	sheet[x3]=dt
	sheet[x4]=tm
	x += 1
wb.save('sent_list')
wb=xl.Workbook()
sheet=wb.active
sheet.title='not sent'
sheet['A1']='Name'
sheet['A2']='email'
x='B'
for name, email in notsent.items():
	sheet[x1]=name
	sheet[x2]=email
	x += 1
wb.save('notsent_list')
smtpObj.quit()



